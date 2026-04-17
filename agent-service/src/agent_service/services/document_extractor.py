from __future__ import annotations

import logging
from typing import Optional

logger = logging.getLogger(__name__)


async def extract_text(
    buffer: bytes,
    mime_type: str,
    file_type: Optional[str] = None,
) -> str:
    """
    Extract plain text from a file buffer.

    Uses ``mime_type`` first; falls back to ``file_type`` (extension) for
    generic mime types. Returns an empty string if extraction is not supported.
    """
    mime = mime_type.lower()
    ext = (file_type or "").lower()

    # PDF
    if mime == "application/pdf" or ext == "pdf":
        return _extract_pdf(buffer)

    # DOCX (Word)
    if mime in (
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/msword",
    ) or ext == "docx":
        return _extract_docx(buffer)

    # XLSX (Excel)
    if mime in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ) or ext == "xlsx":
        return _extract_xlsx(buffer)

    # Plain-text formats
    if (
        mime.startswith("text/")
        or mime in ("application/json", "application/x-ndjson")
        or ext in ("md", "txt", "csv", "json")
    ):
        return buffer.decode("utf-8", errors="replace")

    # Generic octet-stream with known text extensions
    if mime == "application/octet-stream" and ext:
        text_exts = {
            "md", "txt", "csv", "json", "ts", "js", "py", "yaml", "yml",
            "toml", "xml", "html", "css", "sh",
        }
        if ext in text_exts:
            return buffer.decode("utf-8", errors="replace")

    return ""


def _extract_pdf(buffer: bytes) -> str:
    try:
        import fitz  # PyMuPDF

        doc = fitz.open(stream=buffer, filetype="pdf")
        pages: list[str] = []
        for page in doc:
            pages.append(page.get_text())
        doc.close()
        return "\n".join(pages)
    except Exception as exc:
        logger.warning("PDF extraction failed: %s", exc)
        return ""


def _extract_docx(buffer: bytes) -> str:
    try:
        import io

        from docx import Document

        doc = Document(io.BytesIO(buffer))
        return "\n".join(p.text for p in doc.paragraphs)
    except Exception as exc:
        logger.warning("DOCX extraction failed: %s", exc)
        return ""


def _extract_xlsx(buffer: bytes) -> str:
    try:
        import io

        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(buffer), read_only=True, data_only=True)
        parts: list[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[str] = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                rows.append(",".join(cells))
            if rows:
                parts.append(f"## Sheet: {sheet_name}\n" + "\n".join(rows))
        wb.close()
        return "\n\n".join(parts)
    except Exception as exc:
        logger.warning("XLSX extraction failed: %s", exc)
        return ""
